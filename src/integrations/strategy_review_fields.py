#!/usr/bin/env python3
"""Campos da Strategy Review usados pelo Breakeven Auto."""
from __future__ import annotations

import json
import re
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

PROJECT = Path(__file__).resolve().parents[2]
SHEET_ID = "1MrUklD9tulNHsxWmAh3fcUUBlBdY-IHzSUuEPAz32YQ"
SHEET_TAB = "Start Strategy Review"
CRED_PATH = PROJECT.parent / "assessor-pessoal" / "mcp" / "credentials" / "google_sheets_token.json"

LT_SOURCE = "Strategy Review col. H — Life Time (contrato ativo)"
MRR_SOURCE = "Strategy Review col. L — MRR"
# Legado — não usar em manifests novos
TM_RECURRENCE_SOURCE = MRR_SOURCE


def parse_mrr_months(raw: object) -> int | None:
    """Extrai meses de recorrência a partir do texto da col. L (MRR)."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        months = int(raw)
        return months if months > 1 else None
    text = str(raw).strip().lower()
    normalized = "".join(
        c for c in __import__("unicodedata").normalize("NFKD", text) if not __import__("unicodedata").combining(c)
    )
    if not normalized or normalized in {"-", "n/a", "na", "none", "sem", "0", "sem recorrencia"}:
        return None
    if normalized.startswith("sem "):
        return None
    match = re.search(r"(\d+)", text)
    if not match:
        return None
    months = int(match.group(1))
    return months if months > 1 else None


def parse_lt_months(raw: object) -> int | None:
    if raw is None:
        return None
    match = re.search(r"(\d+)", str(raw))
    return int(match.group(1)) if match else None


def resolve_mrr_from_manifest(manifest: dict[str, Any]) -> tuple[int | None, str | None]:
    """Recorrência TM só quando a col. L (MRR) da SR estiver preenchida.

    Ignora `tm_recurrence_months` numérico sem texto MRR — evita confundir LT (col. H) com MRR.
    """
    raw = manifest.get("mrr_raw") or manifest.get("tm_recurrence_raw") or manifest.get("tm_recurrence")
    if raw is None or not str(raw).strip():
        return None, None
    months = manifest.get("mrr_months")
    if months is not None:
        parsed = int(months)
        return (parsed if parsed > 1 else None), str(raw).strip()
    parsed = parse_mrr_months(raw)
    return parsed, str(raw).strip() if parsed else None


def load_credentials() -> Credentials:
    info = json.loads(CRED_PATH.read_text(encoding="utf-8"))
    creds = Credentials(
        token=info.get("token"),
        refresh_token=info.get("refresh_token"),
        token_uri=info.get("token_uri", "https://oauth2.googleapis.com/token"),
        client_id=info.get("client_id"),
        client_secret=info.get("client_secret"),
        scopes=list(info.get("scopes") or []),
    )
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return creds


def read_sr_row(row: int) -> dict[str, Any]:
    """Lê LT (H), fee, mídia, margem, MRR (L) e GrowthPack (N) de uma linha da SR."""
    drive = build("drive", "v3", credentials=load_credentials())
    xlsx = drive.files().export(
        fileId=SHEET_ID,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ).execute()
    wb = openpyxl.load_workbook(BytesIO(xlsx), read_only=True, data_only=True)
    ws = wb[SHEET_TAB] if SHEET_TAB in wb.sheetnames else wb.active
    mrr_raw = ws.cell(row, 12).value
    mrr_text = str(mrr_raw).strip() if mrr_raw is not None and str(mrr_raw).strip() else None
    payload = {
        "row": row,
        "project": ws.cell(row, 2).value,
        "lt_months": parse_lt_months(ws.cell(row, 8).value),
        "mrr_raw": mrr_text,
        "mrr_months": parse_mrr_months(mrr_raw) if mrr_text else None,
    }
    wb.close()
    return payload


def apply_sr_fields_to_manifest(manifest: dict[str, Any], sr_row: int) -> dict[str, Any]:
    """Atualiza manifest com LT (H) e MRR (L) lidos online da SR."""
    sr = read_sr_row(sr_row)
    if sr.get("lt_months"):
        manifest["lt_months"] = sr["lt_months"]
        manifest["lt_source"] = LT_SOURCE
    if sr.get("mrr_months") and sr.get("mrr_raw"):
        manifest["mrr_months"] = sr["mrr_months"]
        manifest["mrr_raw"] = sr["mrr_raw"]
        manifest["mrr_source"] = MRR_SOURCE
        manifest["tm_recurrence_months"] = sr["mrr_months"]
        manifest["tm_recurrence_raw"] = sr["mrr_raw"]
        manifest["tm_recurrence_source"] = MRR_SOURCE
    else:
        for key in (
            "mrr_months",
            "mrr_raw",
            "mrr_source",
            "tm_recurrence_months",
            "tm_recurrence_raw",
            "tm_recurrence_source",
        ):
            manifest.pop(key, None)
    return manifest
