#!/usr/bin/env python3
"""Gera breakeven nativo de inside sales a partir do funil do Growth Pack."""
from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from statistics import median
from typing import Any

import openpyxl
import xlsxwriter


@dataclass
class MonthData:
    label: str
    fee: float
    media: float
    impressions: float
    clicks: float
    leads: float
    mqls: float
    sqls: float
    sales: float
    revenue: float


PROJECT = Path(__file__).resolve().parents[2]


def brl(value: float) -> str:
    text = f"{abs(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{'-' if value < 0 else ''}R$ {text}"


def pct(value: float) -> str:
    return f"{value * 100:.2f}%".replace(".", ",")


def num(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def safe_div(a: float, b: float) -> float:
    return a / b if b else 0.0


def month_label(raw_month: Any, raw_year: Any, raw_date: Any) -> str:
    text = str(raw_date or "")[:7]
    if text and text.lower() != "none":
        return text
    month = str(raw_month or "").strip()
    year = int(num(raw_year))
    if month and month.lower() != "none" and year:
        return f"{month.title()}/{year}"
    return "Mês"


def load_months(project_folder: Path) -> list[MonthData]:
    wb = openpyxl.load_workbook(project_folder / "source" / "growthpack.xlsx", data_only=True)
    ws = wb["6.0 Acompanhamento Mensal"]
    rows = {
        "year": 1,
        "date": 2,
        "month": 4,
        "fee": 5,
        "media": 8,
        "impressions": 13,
        "clicks": 15,
        "leads": 18,
        "mqls": 20,
        "sqls": 21,
        "sales": 25,
        "revenue": 26,
    }
    months: list[MonthData] = []
    for col in range(2, ws.max_column + 1):
        revenue = num(ws.cell(rows["revenue"], col).value)
        sales = num(ws.cell(rows["sales"], col).value)
        media = num(ws.cell(rows["media"], col).value)
        fee = num(ws.cell(rows["fee"], col).value)
        impressions = num(ws.cell(rows["impressions"], col).value)
        clicks = num(ws.cell(rows["clicks"], col).value)
        leads = num(ws.cell(rows["leads"], col).value)
        mqls = num(ws.cell(rows["mqls"], col).value)
        sqls = num(ws.cell(rows["sqls"], col).value)
        # Recorte válido: mês com custo, faturamento e funil inside sales completo.
        if (
            fee <= 0
            or media <= 0
            or sales <= 0
            or revenue <= 0
            or impressions <= 0
            or clicks <= 0
            or leads <= 0
            or mqls <= 0
            or sqls <= 0
        ):
            continue
        months.append(
            MonthData(
                label=month_label(
                    ws.cell(rows["month"], col).value,
                    ws.cell(rows["year"], col).value,
                    ws.cell(rows["date"], col).value,
                ),
                fee=fee,
                media=media,
                impressions=impressions,
                clicks=clicks,
                leads=leads,
                mqls=mqls,
                sqls=sqls,
                sales=sales,
                revenue=revenue,
            )
        )
    if not months:
        raise ValueError("Nenhum mês válido encontrado para funil inside sales.")
    return months


def rates(month: MonthData) -> dict[str, float]:
    return {
        "impression_click": safe_div(month.clicks, month.impressions),
        "click_lead": safe_div(month.leads, month.clicks),
        "lead_mql": safe_div(month.mqls, month.leads),
        "mql_sql": safe_div(month.sqls, month.mqls),
        "sql_sale": safe_div(month.sales, month.sqls),
        "impression_sale": safe_div(month.sales, month.impressions),
    }


def medians(months: list[MonthData]) -> dict[str, float]:
    keys = list(rates(months[0]).keys())
    return {
        key: median([value for value in (rates(month)[key] for month in months) if value > 0] or [0.0])
        for key in keys
    }


def historical_summary(months: list[MonthData], margin: float) -> dict[str, Any]:
    balance = 0.0
    first_breakeven = None
    rows = []
    for month in months:
        cost = month.fee + month.media
        mc = month.revenue * margin
        result = mc - cost
        balance += result
        if first_breakeven is None and balance >= 0:
            first_breakeven = month.label
        rows.append(
            {
                "month": month.label,
                "cost": cost,
                "mc": mc,
                "result": result,
                "cumulative": balance,
            }
        )
    return {
        "fee": sum(month.fee for month in months),
        "media": sum(month.media for month in months),
        "cost": sum(month.fee + month.media for month in months),
        "revenue": sum(month.revenue for month in months),
        "mc": sum(month.revenue for month in months) * margin,
        "result": balance,
        "first_breakeven": first_breakeven,
        "rows": rows,
    }


def build_projection(
    months: list[MonthData],
    margin: float,
    monthly_fee: float,
    monthly_media: float,
    growth: float,
    rate_multiplier: float,
    max_months: int = 24,
) -> list[dict[str, float | str]]:
    current = months[-1]
    ticket = safe_div(sum(m.revenue for m in months), sum(m.sales for m in months))
    bench = medians(months)
    balance = historical_summary(months, margin)["result"]
    rows = []
    revenue = current.revenue
    for idx in range(1, max_months + 1):
        revenue *= 1 + growth
        sales = revenue / ticket if ticket else 0
        sql_rate = min(1.0, bench["sql_sale"] * rate_multiplier) or 0.01
        mql_sql = min(1.0, bench["mql_sql"] * rate_multiplier) or 0.01
        lead_mql = min(1.0, bench["lead_mql"] * rate_multiplier) or 0.01
        click_lead = min(1.0, bench["click_lead"] * rate_multiplier) or 0.01
        impression_click = min(1.0, bench["impression_click"] * rate_multiplier) or 0.001
        sqls = sales / sql_rate
        mqls = sqls / mql_sql
        leads = mqls / lead_mql
        clicks = leads / click_lead
        impressions = clicks / impression_click
        monthly_result = revenue * margin - (monthly_fee + monthly_media)
        balance += monthly_result
        rows.append(
            {
                "month": f"Mês {idx}",
                "fee": monthly_fee,
                "media": monthly_media,
                "impressions": impressions,
                "clicks": clicks,
                "leads": leads,
                "mqls": mqls,
                "sqls": sqls,
                "sales": sales,
                "revenue": revenue,
                "result": monthly_result,
                "cumulative": balance,
                "breakeven": "Sim" if balance >= 0 else "Não",
            }
        )
        if balance >= 0:
            break
    return rows


def write_workbook(path: Path, client: str, months: list[MonthData], margin: float, monthly_fee: float, monthly_media: float) -> dict[str, Any]:
    summary = historical_summary(months, margin)
    scenarios = {
        "Pessimista": build_projection(months, margin, monthly_fee, monthly_media, 0.04, 0.95),
        "Realista": build_projection(months, margin, monthly_fee, monthly_media, 0.10, 1.00),
        "Otimista": build_projection(months, margin, monthly_fee, monthly_media, 0.16, 1.08),
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(path)
    workbook.set_properties({"title": f"Breakeven Inside Sales — {client}", "author": "QGI"})
    money = workbook.add_format({"num_format": 'R$ #,##0.00;[Red]-R$ #,##0.00', "border": 1})
    number = workbook.add_format({"num_format": "#,##0", "border": 1})
    pct_fmt = workbook.add_format({"num_format": "0.00%", "border": 1})
    normal = workbook.add_format({"border": 1})
    header = workbook.add_format({"bold": True, "bg_color": "#102A43", "font_color": "#FFFFFF", "border": 1})
    title = workbook.add_format({"bold": True, "font_size": 16, "bg_color": "#102A43", "font_color": "#FFFFFF"})
    good = workbook.add_format({"bold": True, "bg_color": "#D9EAD3", "font_color": "#38761D", "border": 1})
    bad = workbook.add_format({"bold": True, "bg_color": "#F4CCCC", "font_color": "#990000", "border": 1})

    ws = workbook.add_worksheet("Resumo")
    ws.set_column("A:A", 34)
    ws.set_column("B:B", 24)
    ws.merge_range("A1:B1", f"Breakeven Inside Sales — {client}", title)
    rows = [
        ("Modelo", "Inside sales nativo do Growth Pack", normal),
        ("Funil", "Impressões → Cliques → Leads → MQLs → SQLs → Vendas → Faturamento", normal),
        ("Período histórico", f"{months[0].label} a {months[-1].label}", normal),
        ("Margem usada", margin, pct_fmt),
        ("Fee mensal futuro", monthly_fee, money),
        ("Mídia mensal futura", monthly_media, money),
        ("Breakeven da competência", (monthly_fee + monthly_media) / margin, money),
        ("Resultado histórico acumulado", summary["result"], money),
        ("Status histórico", "Breakeven já atingido" if summary["result"] >= 0 else "Breakeven ainda não atingido", good if summary["result"] >= 0 else bad),
        ("Primeiro mês breakeven histórico", summary["first_breakeven"] or "Não atingiu no histórico", normal),
    ]
    for idx, (label, value, fmt) in enumerate(rows, 3):
        ws.write(idx, 0, label, header if idx == 3 else normal)
        if isinstance(value, (int, float)):
            ws.write_number(idx, 1, value, fmt)
        else:
            ws.write(idx, 1, value, fmt)

    hist = workbook.add_worksheet("Histórico GP")
    hist.set_column("A:N", 16)
    headers = ["Mês", "Fee", "Mídia", "Impressões", "Cliques", "Leads", "MQLs", "SQLs", "Vendas", "Faturamento", "MC", "Resultado mês", "Resultado acum.", "Status"]
    for col, text in enumerate(headers):
        hist.write(0, col, text, header)
    cumulative = 0.0
    for row, month in enumerate(months, 1):
        cost = month.fee + month.media
        mc = month.revenue * margin
        result = mc - cost
        cumulative += result
        values = [month.label, month.fee, month.media, month.impressions, month.clicks, month.leads, month.mqls, month.sqls, month.sales, month.revenue, mc, result, cumulative, "Breakeven" if cumulative >= 0 else "Déficit"]
        for col, value in enumerate(values):
            fmt = money if col in (1, 2, 9, 10, 11, 12) else number if col in (3, 4, 5, 6, 7, 8) else good if col == 13 and value == "Breakeven" else bad if col == 13 else normal
            if isinstance(value, (int, float)):
                hist.write_number(row, col, value, fmt)
            else:
                hist.write(row, col, value, fmt)

    bench = workbook.add_worksheet("Bench Funil")
    bench.set_column("A:H", 20)
    rate_headers = ["Mês", "Imp. → Clique", "Clique → Lead", "Lead → MQL", "MQL → SQL", "SQL → Venda", "Imp. → Venda"]
    for col, text in enumerate(rate_headers):
        bench.write(0, col, text, header)
    for row, month in enumerate(months, 1):
        row_rates = rates(month)
        vals = [month.label, row_rates["impression_click"], row_rates["click_lead"], row_rates["lead_mql"], row_rates["mql_sql"], row_rates["sql_sale"], row_rates["impression_sale"]]
        for col, value in enumerate(vals):
            if isinstance(value, float):
                bench.write_number(row, col, value, pct_fmt)
            else:
                bench.write(row, col, value, normal)
    med = medians(months)
    median_row = len(months) + 2
    bench.write(median_row, 0, "MEDIANA", header)
    for col, key in enumerate(["impression_click", "click_lead", "lead_mql", "mql_sql", "sql_sale", "impression_sale"], 1):
        bench.write_number(median_row, col, med[key], pct_fmt)

    for name, projection in scenarios.items():
        sh = workbook.add_worksheet(name)
        sh.set_column("A:M", 16)
        headers = ["Mês", "Fee", "Mídia", "Impressões", "Cliques", "Leads", "MQLs", "SQLs", "Vendas", "Faturamento", "Resultado mês", "Resultado acum.", "Breakeven?"]
        for col, text in enumerate(headers):
            sh.write(0, col, text, header)
        for row, item in enumerate(projection, 1):
            values = [item["month"], item["fee"], item["media"], item["impressions"], item["clicks"], item["leads"], item["mqls"], item["sqls"], item["sales"], item["revenue"], item["result"], item["cumulative"], item["breakeven"]]
            for col, value in enumerate(values):
                fmt = money if col in (1, 2, 9, 10, 11) else number if col in (3, 4, 5, 6, 7, 8) else good if col == 12 and value == "Sim" else bad if col == 12 else normal
                if isinstance(value, (int, float)):
                    sh.write_number(row, col, value, fmt)
                else:
                    sh.write(row, col, value, fmt)

    prem = workbook.add_worksheet("Premissas")
    prem.set_column("A:B", 32)
    prem.write("A1", "Premissa", header)
    prem.write("B1", "Valor", header)
    for row, (key, value) in enumerate(
        [
            ("Fonte", "Growth Pack > 6.0 Acompanhamento Mensal"),
            ("Fee/mídia/margem futura", "Flow/Strategy Review"),
            ("Horizonte máximo", "24 meses por cenário"),
            ("Critério de parada", "parar quando resultado acumulado >= 0"),
            ("Pessimista", "crescimento de faturamento 4% a.m.; taxas 95% do bench"),
            ("Realista", "crescimento de faturamento 10% a.m.; taxas no bench"),
            ("Otimista", "crescimento de faturamento 16% a.m.; taxas 108% do bench"),
        ],
        1,
    ):
        prem.write(row, 0, key, normal)
        prem.write(row, 1, value, normal)

    workbook.close()
    return {"summary": summary, "scenarios": scenarios}


def write_report(path: Path, client: str, months: list[MonthData], margin: float, monthly_fee: float, monthly_media: float, result: dict[str, Any]) -> None:
    summary = result["summary"]
    scenario_lines = []
    for name, rows in result["scenarios"].items():
        last = rows[-1]
        scenario_lines.append(
            f"| {name} | {len(rows)} | {brl(last['revenue'])} | {brl(last['cumulative'])} | {'Sim' if last['cumulative'] >= 0 else 'Não no horizonte'} |"
        )
    lines = [
        f"# [Gerência] - [Análise e Estratégia] - [{client}]",
        "",
        "## Contexto",
        "",
        "Análise refeita com o funil inside sales nativo encontrado no Growth Pack.",
        "",
        "Funil usado: **impressões → cliques → leads → MQLs → SQLs → vendas → faturamento**.",
        "",
        "## Breakeven histórico",
        "",
        f"- Período histórico válido: {months[0].label} a {months[-1].label}.",
        f"- Resultado histórico acumulado: **{brl(summary['result'])}**.",
        f"- Status histórico: **{'breakeven já atingido' if summary['result'] >= 0 else 'breakeven ainda não atingido'}**.",
        f"- Primeiro mês positivo no acumulado: {summary['first_breakeven'] or 'não atingiu no histórico'}.",
        "",
        "## Premissas financeiras futuras",
        "",
        f"- Fee mensal: {brl(monthly_fee)}.",
        f"- Mídia mensal: {brl(monthly_media)}.",
        f"- Margem de contribuição: {pct(margin)}.",
        f"- Breakeven da competência: {brl((monthly_fee + monthly_media) / margin)}.",
        "",
        "## Cenários com horizonte variável",
        "",
        "| Cenário | Meses projetados | Faturamento no último mês | Saldo final | Breakeva? |",
        "|---|---:|---:|---:|---|",
        *scenario_lines,
        "",
        "## Observações",
        "",
        "- A projeção não está travada em 7 meses; cada cenário segue até breakevar ou até o limite de 24 meses.",
        "- E-commerce permanece com o funil padrão da skill; inside sales deve sempre usar o funil real do Growth Pack.",
        "- Esta versão substitui a adaptação anterior que forçava etapas de e-commerce na Soma.",
        "",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera planilha e relatório inside sales nativo.")
    parser.add_argument("--project-folder", type=Path, required=True)
    args = parser.parse_args()
    project_folder = args.project_folder
    manifest = json.loads((project_folder / "source" / "manifest-entry.json").read_text(encoding="utf-8"))
    client = manifest["name"]
    margin = num(manifest["margin_pct"]) / 100
    monthly_fee = num(manifest["fee"])
    monthly_media = num(manifest["media_planned"])
    months = load_months(project_folder)
    spreadsheet = project_folder / "spreadsheet" / "Soma Soluções - Breakeven Inside Sales Nativo.xlsx"
    report = project_folder / "report" / "[Gerência] - [Análise e Estratégia] - [Soma Soluções - Inside Sales Nativo].md"
    result = write_workbook(spreadsheet, client, months, margin, monthly_fee, monthly_media)
    write_report(report, client, months, margin, monthly_fee, monthly_media, result)
    print(spreadsheet)
    print(report)


if __name__ == "__main__":
    main()
