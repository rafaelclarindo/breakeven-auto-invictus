#!/usr/bin/env python3
"""Lê Growth Packs direto do Google Sheets (sem download .xlsx)."""
from __future__ import annotations

import io
import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

PROJECT = Path(__file__).resolve().parents[2]
CRED_PATH = PROJECT.parent / "assessor-pessoal" / "mcp" / "credentials" / "google_sheets_token.json"
OAUTH_ACCOUNT_EMAIL = "rafael.clarindo@v4company.com"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

EXCEL_EPOCH = datetime(1899, 12, 30)
SERIAL_DATE_MIN = 30_000  # ~1982
SERIAL_DATE_MAX = 60_000  # ~2064


class SheetCell:
    __slots__ = ("value",)

    def __init__(self, value: Any) -> None:
        self.value = value


class SheetWorksheetAdapter:
    """Interface mínima compatível com openpyxl worksheet (.cell, .max_column)."""

    def __init__(self, grid: list[list[Any]]) -> None:
        self._grid = grid
        self.max_column = max((len(row) for row in grid), default=0)

    def cell(self, row: int, column: int) -> SheetCell:
        if row < 1 or column < 1:
            return SheetCell(None)
        row_idx = row - 1
        col_idx = column - 1
        if row_idx >= len(self._grid):
            return SheetCell(None)
        row_values = self._grid[row_idx]
        if col_idx >= len(row_values):
            return SheetCell(None)
        return SheetCell(row_values[col_idx])


class SheetGrid:
    def __init__(self, grid: list[list[Any]]) -> None:
        self._grid = grid

    def as_worksheet(self) -> SheetWorksheetAdapter:
        return SheetWorksheetAdapter(self._grid)


def load_google_credentials() -> Credentials:
    if not CRED_PATH.exists():
        raise FileNotFoundError(f"OAuth token não encontrado: {CRED_PATH}")
    info = json.loads(CRED_PATH.read_text(encoding="utf-8"))
    creds = Credentials(
        token=info.get("token"),
        refresh_token=info.get("refresh_token"),
        token_uri=info.get("token_uri", "https://oauth2.googleapis.com/token"),
        client_id=info.get("client_id"),
        client_secret=info.get("client_secret"),
        scopes=list(info.get("scopes") or []),
    )
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
        info.update(
            {
                "token": creds.token,
                "expiry": creds.expiry.isoformat() if creds.expiry else None,
            }
        )
        CRED_PATH.write_text(json.dumps(info, ensure_ascii=False, indent=2), encoding="utf-8")
    return creds


def extract_spreadsheet_id(link: str | None) -> str | None:
    if not link:
        return None
    patterns = [
        r"/spreadsheets/d/([^/?#]+)",
        r"/file/d/([^/?#]+)",
        r"[?&]id=([^&#]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, link)
        if match:
            return match.group(1)
    return None


def _serial_to_datetime(serial: float) -> datetime:
    return EXCEL_EPOCH + timedelta(days=float(serial))


def _coerce_grid_value(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        serial = float(value)
        if SERIAL_DATE_MIN <= serial <= SERIAL_DATE_MAX:
            return _serial_to_datetime(serial)
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        try:
            serial = float(text.replace(",", "."))
        except ValueError:
            return value
        if SERIAL_DATE_MIN <= serial <= SERIAL_DATE_MAX:
            return _serial_to_datetime(serial)
        return value
    return value


def _normalize_grid(raw_rows: list[list[Any]] | None) -> list[list[Any]]:
    if not raw_rows:
        return []
    width = max(len(row) for row in raw_rows)
    grid: list[list[Any]] = []
    for row in raw_rows:
        padded = [_coerce_grid_value(cell) for cell in row]
        if len(padded) < width:
            padded.extend([None] * (width - len(padded)))
        grid.append(padded)
    return grid


def read_sheet_grid(
    creds: Credentials,
    spreadsheet_id: str,
    sheet_name: str,
    *,
    max_rows: int = 200,
    max_cols: int = 120,
) -> SheetGrid:
    """Busca aba via Sheets API (values.get, UNFORMATTED_VALUE + serial dates)."""
    sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)
    end_col = _column_letter(max_cols)
    range_a1 = f"'{sheet_name}'!A1:{end_col}{max_rows}"
    response = (
        sheets.spreadsheets()
        .values()
        .get(
            spreadsheetId=spreadsheet_id,
            range=range_a1,
            valueRenderOption="UNFORMATTED_VALUE",
            dateTimeRenderOption="SERIAL_NUMBER",
        )
        .execute()
    )
    return SheetGrid(_normalize_grid(response.get("values")))


def _sheets_api_disabled(error: HttpError) -> bool:
    if error.resp.status != 403:
        return False
    payload = error.error_details if hasattr(error, "error_details") else []
    for detail in payload or []:
        if detail.get("reason") == "SERVICE_DISABLED":
            return True
    return "SERVICE_DISABLED" in str(error)


def read_workbook_sheet_via_drive_export(
    creds: Credentials,
    spreadsheet_id: str,
    sheet_name: str,
) -> Any:
    """Fallback: exporta XLSX na memória via Drive API (sem gravar em disco)."""
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)
    try:
        payload = drive.files().export(fileId=spreadsheet_id, mimeType=XLSX_MIME).execute()
    except HttpError as error:
        raise RuntimeError(
            f"Drive export falhou para {spreadsheet_id}: HTTP {error.resp.status}. "
            f"Compartilhe o GP com a conta OAuth ({OAUTH_ACCOUNT_EMAIL})."
        ) from error
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(
            f"Aba {sheet_name!r} ausente no GP exportado. Abas: {workbook.sheetnames}"
        )
    return workbook[sheet_name]


def list_spreadsheet_sheet_titles(creds: Credentials, spreadsheet_id: str) -> list[str]:
    sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)
    meta = sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id, fields="sheets.properties.title"
    ).execute()
    return [item["properties"]["title"] for item in meta.get("sheets", [])]


def find_acompanhamento_mensal_sheet(creds: Credentials, spreadsheet_id: str) -> str:
    """Retorna a aba de acompanhamento mensal do GP (6.0 preferencial, depois 2.2)."""
    titles = list_spreadsheet_sheet_titles(creds, spreadsheet_id)
    matches = [
        title
        for title in titles
        if "acompanhamento mensal" in str(title).casefold()
        or str(title).casefold().startswith("acomp. mensal")
    ]
    if not matches:
        raise KeyError(
            f"Nenhuma aba 'Acompanhamento Mensal' em {spreadsheet_id}. "
            f"Abas: {titles}"
        )
    for prefix in ("6.0", "6.", "2.2", "2."):
        for title in matches:
            if title.startswith(prefix):
                return title
    return matches[0]


def _column_letter(index: int) -> str:
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters or "A"


def open_growthpack_worksheet(
    project_folder: Path,
    sheet_name: str,
    *,
    source: str = "online",
) -> tuple[Any, str]:
    """Abre aba do Growth Pack. Default: leitura online (Sheets API).

    Returns:
        (worksheet, source_used) where source_used is ``online`` or ``local``.
    """
    manifest_path = project_folder / "source" / "manifest-entry.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"Manifest não encontrado: {manifest_path}")

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    local_path = project_folder / "source" / "growthpack.xlsx"
    spreadsheet_id = extract_spreadsheet_id(manifest.get("growthpack_updated_link"))

    if source not in {"online", "local", "auto"}:
        raise ValueError(f"source inválido: {source!r} (use online, local ou auto)")

    if source in {"online", "auto"}:
        if not spreadsheet_id:
            if source == "online":
                raise ValueError(
                    f"Link do Growth Pack ausente em {manifest_path} "
                    "(growthpack_updated_link)."
                )
        else:
            creds = load_google_credentials()
            online_errors: list[str] = []
            try:
                grid = read_sheet_grid(creds, spreadsheet_id, sheet_name)
                return grid.as_worksheet(), "online"
            except HttpError as error:
                online_errors.append(f"Sheets API HTTP {error.resp.status}")
                if _sheets_api_disabled(error):
                    online_errors.append(
                        "Sheets API desabilitada no GCP — tentando Drive export em memória"
                    )
            try:
                worksheet = read_workbook_sheet_via_drive_export(
                    creds, spreadsheet_id, sheet_name
                )
                return worksheet, "online-drive-export"
            except (HttpError, RuntimeError, KeyError) as error:
                online_errors.append(str(error))
                if source == "online":
                    raise RuntimeError(
                        f"Leitura online falhou para {spreadsheet_id}. "
                        + " | ".join(online_errors)
                        + f". Compartilhe o GP com {OAUTH_ACCOUNT_EMAIL} "
                        "ou use --gp-source local."
                    ) from error

    if source in {"local", "auto"}:
        if local_path.exists():
            workbook = openpyxl.load_workbook(local_path, data_only=True)
            if sheet_name not in workbook.sheetnames:
                raise KeyError(
                    f"Aba {sheet_name!r} ausente em {local_path}. "
                    f"Abas: {workbook.sheetnames}"
                )
            return workbook[sheet_name], "local"

    hint = (
        f"Compartilhe o GP com a conta OAuth ou coloque {local_path.name} manualmente."
    )
    if spreadsheet_id:
        raise FileNotFoundError(
            f"Growth Pack indisponível (online e local). ID: {spreadsheet_id}. {hint}"
        )
    raise FileNotFoundError(f"Growth Pack indisponível. {hint}")
