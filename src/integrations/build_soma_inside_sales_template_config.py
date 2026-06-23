#!/usr/bin/env python3
"""Monta config da Soma para o template da skill com funil inside sales."""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from statistics import median
from typing import Any

import openpyxl

MAX_CONVERSION_RATE = 0.95
MIN_COST_PER_IMPRESSION = 0.01


def num(value: Any) -> float:
    try:
        return float(value) if value is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def div(a: float, b: float) -> float:
    return a / b if b else 0.0


def cap_rate(value: float) -> float:
    return min(MAX_CONVERSION_RATE, value)


def gradual(start: float, end: float, months: int = 7) -> list[float]:
    return [cap_rate(start + (end - start) * idx / (months - 1)) for idx in range(months)]


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera config Soma inside sales para template da skill.")
    parser.add_argument("--project-folder", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    project_folder = args.project_folder
    manifest = json.loads((project_folder / "source" / "manifest-entry.json").read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(project_folder / "source" / "growthpack.xlsx", data_only=True)
    ws = wb["6.0 Acompanhamento Mensal"]
    rows = {
        "date": 2,
        "fee": 5,
        "media": 8,
        "impressions": 13,
        "clicks": 15,
        "leads": 18,
        "lead_quali": 19,
        "mqls": 20,
        "sqls": 21,
        "sales": 25,
        "revenue": 26,
    }
    months = []
    for col in range(2, ws.max_column + 1):
        item = {
            "month": str(ws.cell(rows["date"], col).value or "")[:7],
            "fee": num(ws.cell(rows["fee"], col).value),
            "media": num(ws.cell(rows["media"], col).value),
            "impressions": num(ws.cell(rows["impressions"], col).value),
            "clicks": num(ws.cell(rows["clicks"], col).value),
            "leads": num(ws.cell(rows["leads"], col).value),
            "lead_quali": num(ws.cell(rows["lead_quali"], col).value),
            "mqls": num(ws.cell(rows["mqls"], col).value),
            "sqls": num(ws.cell(rows["sqls"], col).value),
            "sales": num(ws.cell(rows["sales"], col).value),
            "revenue": num(ws.cell(rows["revenue"], col).value),
        }
        if all(
            item[key] > 0
            for key in (
                "fee",
                "media",
                "impressions",
                "clicks",
                "leads",
                "lead_quali",
                "mqls",
                "sqls",
                "sales",
                "revenue",
            )
        ):
            months.append(item)
    if not months:
        raise ValueError("Nenhum mês com funil completo encontrado.")

    margin = num(manifest["margin_pct"]) / 100
    monthly_fee = num(manifest["fee"])
    monthly_media = num(manifest["media_planned"])
    ticket = div(sum(m["revenue"] for m in months), sum(m["sales"] for m in months))
    current = months[-1]
    current_cps = div(current["media"], current["impressions"])
    min_cps = max(MIN_COST_PER_IMPRESSION, current_cps * 0.85)

    rates = {
        "impression_click": median(div(m["clicks"], m["impressions"]) for m in months),
        "click_lead": median(div(m["leads"], m["clicks"]) for m in months),
        "lead_lead_quali": median(div(m["lead_quali"], m["leads"]) for m in months),
        "lead_quali_mql": median(div(m["mqls"], m["lead_quali"]) for m in months),
        "mql_sql": median(div(m["sqls"], m["mqls"]) for m in months),
        "sql_sale": median(div(m["sales"], m["sqls"]) for m in months),
    }
    current_rates = {
        "impression_click": div(current["clicks"], current["impressions"]),
        "click_lead": div(current["leads"], current["clicks"]),
        "lead_lead_quali": div(current["lead_quali"], current["leads"]),
        "lead_quali_mql": div(current["mqls"], current["lead_quali"]),
        "mql_sql": div(current["sqls"], current["mqls"]),
        "sql_sale": div(current["sales"], current["sqls"]),
    }

    def scenario_end(current_val: float, bench_val: float, multiplier: float) -> float:
        """Meta de taxa no horizonte — nunca piora o Mês 1 (sempre = atual)."""
        if multiplier < 1.0:
            return cap_rate(min(current_val * 0.98, bench_val * multiplier))
        return cap_rate(max(current_val * multiplier, bench_val * multiplier))

    def scenario_rate_series(rate_key: str, multiplier: float) -> list[float]:
        current_val = current_rates[rate_key]
        end_val = scenario_end(current_val, rates[rate_key], multiplier)
        return gradual(current_val, end_val)

    source_months = [
        [m["month"], m["fee"], m["media"], m["impressions"], m["sqls"], m["sales"], m["revenue"]]
        for m in months
    ]
    benchmark_months = [
        [
            m["month"],
            m["impressions"],
            m["clicks"],
            m["leads"],
            m["lead_quali"],
            m["mqls"],
            m["sqls"],
            m["sales"],
            m["sales"],
        ]
        for m in months
    ]
    current_funnel = {
        "sessions": current["impressions"],
        "page_view": current["impressions"],
        "view_item": current["clicks"],
        "add_to_cart": current["leads"],
        "view_cart": current["lead_quali"],
        "begin_checkout": current["mqls"],
        "add_shipping_info": current["sqls"],
        "add_payment_info": current["sales"],
        "orders": current["sqls"],
        "sales": current["sales"],
        "purchase": current["sales"],
        "revenue": current["revenue"],
        "media": current["media"],
    }

    def scenario(name: str, growth: float, multiplier: float, color: str) -> dict:
        revenue = []
        value = current["revenue"]
        for _ in range(7):
            value *= 1 + growth
            revenue.append(round(value, 2))
        imp_click = scenario_rate_series("impression_click", multiplier)
        click_lead = scenario_rate_series("click_lead", multiplier)
        lead_lead_quali = scenario_rate_series("lead_lead_quali", multiplier)
        lead_quali_mql = scenario_rate_series("lead_quali_mql", multiplier)
        mql_sql = scenario_rate_series("mql_sql", multiplier)
        sql_sale = scenario_rate_series("sql_sale", multiplier)
        return {
            "media": [monthly_media] * 7,
            "revenue": revenue,
            "ticket": [round(ticket * (1 + 0.005 * idx), 2) for idx in range(7)],
            "session_view": imp_click,
            "view_add": click_lead,
            "view_cart_share": [
                imp_click[i] * click_lead[i] * lead_lead_quali[i] * lead_quali_mql[i] for i in range(7)
            ],
            "add_view_cart": lead_lead_quali,
            "viewcart_checkout": lead_quali_mql,
            "checkout_shipping": mql_sql,
            "shipping_payment": sql_sale,
            "payment_order": [1.0] * 7,
            "order_sale": [1.0] * 7,
            "tab_color": color,
        }

    realista_scenario = scenario("Realista", 0.10, 1.05, "#5B9BD5")
    config = {
        "client": manifest["name"],
        "project_model": "Inside Sales",
        "funnel_has_lead_quali": True,
        "projection_rules": {
            "max_conversion_rate": MAX_CONVERSION_RATE,
            "min_cost_per_impression": min_cps,
            "media_lever_after_monthly_breakeven": True,
        },
        "current_period": f"{current['month']} fechado",
        "lt_period": f"{months[0]['month']} a {months[-1]['month']}",
        "margin": margin,
        "monthly_fee": monthly_fee,
        "monthly_media": monthly_media,
        "source_mapping": {
            "fee": "Growth Pack > 6.0 Acompanhamento Mensal > linha 5",
            "media": "Growth Pack > 6.0 Acompanhamento Mensal > linha 8",
            "revenue": "Growth Pack > 6.0 Acompanhamento Mensal > linha 26",
            "funnel": (
                "Impressões linha 13, Cliques linha 15, Leads linha 18, "
                "Lead quali linha 19, MQLs linha 20, SQLs linha 21, Vendas linha 25"
            ),
        },
        "source_months": source_months,
        "benchmark_months": benchmark_months,
        "current_funnel": current_funnel,
        "minimum_scenario": {
            "revenue": list(realista_scenario["revenue"]),
            "session_view": gradual(
                current_rates["impression_click"],
                cap_rate(max(current_rates["impression_click"] * 1.05, rates["impression_click"] * 1.05)),
            ),
            "view_add": gradual(
                current_rates["click_lead"],
                cap_rate(max(current_rates["click_lead"] * 1.05, rates["click_lead"] * 1.05)),
            ),
            "lead_lead_quali": gradual(
                current_rates["lead_lead_quali"],
                cap_rate(current_rates["lead_lead_quali"] * 1.03),
            ),
            "lead_quali_mql": gradual(
                current_rates["lead_quali_mql"],
                cap_rate(current_rates["lead_quali_mql"] * 1.08),
            ),
            "mql_sql": gradual(
                current_rates["mql_sql"],
                cap_rate(max(current_rates["mql_sql"] * 1.05, rates["mql_sql"] * 1.05)),
            ),
            "sql_sale": gradual(
                current_rates["sql_sale"],
                cap_rate(current_rates["sql_sale"] * 1.08),
            ),
            "add_cart_purchase": gradual(
                div(current["sales"], current["leads"]),
                cap_rate(div(current["sales"], current["leads"]) * 1.05),
            ),
            "approval_target": 1.0,
            "order_sale": [1.0] * 7,
        },
        "scenarios": {
            "Pessimista": scenario("Pessimista", 0.04, 0.96, "#C55A11"),
            "Realista": realista_scenario,
            "Otimista": scenario("Otimista", 0.16, 1.12, "#70AD47"),
        },
        "context": {
            "product": "Executar",
            "phase": "Breakeven inside sales com Lead quali nativo no template da skill",
            "main_risk": "Projeto ainda não atingiu breakeven histórico no recorte Jan/2026-Jun/2026.",
            "diagnosis": [
                "Funil real do Growth Pack: impressões, cliques, leads, lead quali, MQLs, SQLs, vendas.",
                "Lead quali é a etapa operacional da Soma — o time comercial não trabalha leads brutos.",
                "Alavanca de mídia extra só entra após resultado líquido mensal positivo; taxas capadas em 95%.",
            ],
            "actions": [
                "Validar com gestão se o horizonte deve ser expandido além das 7 colunas do template",
                "Acompanhar evolução lead quali → MQL → SQL → vendas e faturamento por venda",
            ],
        },
    }
    output = args.output or project_folder / "config-inside-sales-template.json"
    output.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output)


if __name__ == "__main__":
    main()
