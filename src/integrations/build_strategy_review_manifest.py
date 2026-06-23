#!/usr/bin/env python3
"""Gera manifest Strategy Review: ordem col B + insumos Flow (GrowthPack Atualizado)."""
from __future__ import annotations

import json
import os
import re
import sys
import time
import unicodedata
from datetime import date
from io import BytesIO
from pathlib import Path

PROJECT = Path(__file__).resolve().parents[2]
MONITOR = PROJECT.parent / "monitor-invictus"
TATICO = PROJECT.parent / "tatico-gerencial" / "src" / "sistema-invictus"
sys.path.insert(0, str(MONITOR))

SHEET_ID = "1MrUklD9tulNHsxWmAh3fcUUBlBdY-IHzSUuEPAz32YQ"
SHEET_TAB = "Start Strategy Review"
OUT_DIR = PROJECT / "assets"

COORD_DOC_IDS = {
    "jefferson.vieira": "l21p7mxykrlps3xl114c0op8",
    "ueliton.pereira": "xpucxsbd5nbpx220fswil9hq",
    "melissa.pessoa": "ah9hs9t1z6y5avndxdeaf8vx",
    "thaina.gomes": "rvynya430fajpd76o0b43jzt",
    "thiago.almeida": "z8tl6w5co71m6ot1dqpwbkqe",
}

FLOW_COLS = [
    "name",
    "fee",
    "campaigns_budget_milestone_total_qty",
    "results_contribution_margin_pct",
    "paid_traffic_growthpack_updated_link",
]
CLIENT_JSON = TATICO / "prisma" / "data" / "clientExecutar.json"
CRED_PATH = PROJECT.parent / "assessor-pessoal" / "mcp" / "credentials" / "google_sheets_token.json"


def load_env() -> None:
    for env_path in (MONITOR / ".env", TATICO / ".env"):
        if not env_path.exists():
            continue
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
    os.environ.setdefault("COCKPIT_MCP_BEARER", os.environ.get("COCKPIT_TOKEN", ""))
    os.environ.setdefault(
        "COCKPIT_MCP_URL",
        os.environ.get("COCKPIT_URL", "https://mcp-cockpit.dados.collieassociados.com/mcp"),
    )


def norm(s: str | None) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.strip().lower())


def resolve_val(v):
    if v is None:
        return None
    if isinstance(v, dict) and "value" in v:
        return v.get("value")
    return v


def to_num(v):
    v = resolve_val(v)
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def to_str(v):
    v = resolve_val(v)
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def field(row: dict, key: str):
    hst = row.get("healthScoreTable") or {}
    return row.get(key) if row.get(key) is not None else hst.get(key)


def format_margin(v: float | None) -> float | None:
    if v is None:
        return None
    if 0 < v <= 1:
        return round(v * 100, 2)
    return round(v, 2)


def parse_row(row: dict) -> dict | None:
    doc = row.get("documentId")
    if not doc:
        return None
    link = to_str(field(row, "paid_traffic_growthpack_updated_link"))
    if link and not link.startswith("http"):
        link = None
    return {
        "document_id": doc,
        "name": row.get("name") or to_str(field(row, "name")),
        "fee": to_num(field(row, "fee")),
        "media_planned": to_num(field(row, "campaigns_budget_milestone_total_qty")),
        "margin_pct": format_margin(to_num(field(row, "results_contribution_margin_pct"))),
        "growthpack_updated_link": link,
    }


def fetch_flow() -> dict[str, dict]:
    from ingest.lib.mcp_client import cockpit_client

    client = cockpit_client()
    by_doc: dict[str, dict] = {}

    def query_table(payload: dict, *, timeout: float = 120.0, attempts: int = 3):
        last_error: Exception | None = None
        for attempt in range(1, attempts + 1):
            try:
                return client.call_json("cockpit_query_table", payload, timeout=timeout)
            except Exception as exc:  # noqa: BLE001 - surface final MCP failure with context
                last_error = exc
                if attempt == attempts:
                    break
                time.sleep(2 * attempt)
        raise RuntimeError(f"Falha ao consultar Cockpit após {attempts} tentativas") from last_error

    def ingest(rows):
        for row in rows or []:
            parsed = parse_row(row)
            if parsed:
                by_doc[parsed["document_id"]] = parsed

    data = query_table(
        {
            "fetchAllPages": True,
            "filterByCategory": "all",
            "filterByStatus": "all",
            "pageSize": 250,
            "resolveCalculations": True,
            "calculationTimezone": "America/Sao_Paulo",
            "columns": FLOW_COLS,
        }
    )
    rows = data if isinstance(data, list) else data.get("data", [])
    ingest(rows)

    for user_doc in COORD_DOC_IDS.values():
        data = query_table(
            {
                "filterByUser": user_doc,
                "filterByStatus": "active",
                "fetchAllPages": True,
                "pageSize": 250,
                "resolveCalculations": True,
                "calculationTimezone": "America/Sao_Paulo",
                "columns": FLOW_COLS,
            }
        )
        rows = data if isinstance(data, list) else data.get("data", [])
        ingest(rows)

    return by_doc


def build_name_index(by_doc: dict[str, dict]) -> dict[str, dict]:
    clients = {}
    if CLIENT_JSON.exists():
        for r in json.loads(CLIENT_JSON.read_text(encoding="utf-8")):
            if r.get("gerencia") == "invictus" and r.get("active") and r.get("cockpitDocumentId"):
                clients[r["cockpitDocumentId"]] = r

    idx: dict[str, dict] = {}
    for doc, c in clients.items():
        row = by_doc.get(doc, {})
        payload = {
            "document_id": doc,
            "coordinator": c.get("coordinator"),
            "slug": c.get("slug"),
            "fee": row.get("fee"),
            "media_planned": row.get("media_planned"),
            "margin_pct": row.get("margin_pct"),
            "growthpack_updated_link": row.get("growthpack_updated_link"),
        }
        for candidate in (row.get("name"), c.get("name")):
            nk = norm(candidate)
            if nk and nk not in idx:
                idx[nk] = payload
    return idx


def lookup(name_idx: dict[str, dict], project: str) -> dict | None:
    data = name_idx.get(norm(project))
    if data:
        return data
    np = norm(project)
    for nk, payload in name_idx.items():
        if nk in np or np in nk:
            return payload
    return None


def read_sheet_projects() -> list[str]:
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    import openpyxl

    info = json.loads(CRED_PATH.read_text(encoding="utf-8"))
    creds = Credentials(
        token=info.get("token"),
        refresh_token=info.get("refresh_token"),
        token_uri=info.get("token_uri", "https://oauth2.googleapis.com/token"),
        client_id=info.get("client_id"),
        client_secret=info.get("client_secret"),
        scopes=list(info.get("scopes") or []),
    )
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())

    drive = build("drive", "v3", credentials=creds)
    xlsx = drive.files().export(
        fileId=SHEET_ID,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ).execute()
    wb = openpyxl.load_workbook(BytesIO(xlsx), read_only=True, data_only=True)
    ws = wb[SHEET_TAB] if SHEET_TAB in wb.sheetnames else wb.active

    projects: list[str] = []
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        val = row[0] if row else None
        if not val:
            continue
        project = str(val).strip()
        if project.startswith("---") or project.upper() in {"TOTAL", "PROJETO"}:
            continue
        projects.append(project)
    wb.close()
    return projects


def main() -> None:
    load_env()
    today = date.today().isoformat()
    by_doc = fetch_flow()
    name_idx = build_name_index(by_doc)
    sheet_projects = read_sheet_projects()

    projects = []
    for i, name in enumerate(sheet_projects, 1):
        meta = lookup(name_idx, name) or {}
        projects.append({
            "order": i,
            "name": name,
            "document_id": meta.get("document_id"),
            "coordinator": meta.get("coordinator"),
            "slug": meta.get("slug"),
            "fee": meta.get("fee"),
            "media_planned": meta.get("media_planned"),
            "margin_pct": meta.get("margin_pct"),
            "growthpack_updated_link": meta.get("growthpack_updated_link"),
            "cockpit_fields": {
                "fee": "fee",
                "media_planned": "campaigns_budget_milestone_total_qty",
                "margin_pct": "results_contribution_margin_pct",
                "growthpack_updated_link": "paid_traffic_growthpack_updated_link",
            },
        })

    manifest = {
        "sheet_id": SHEET_ID,
        "sheet_tab": SHEET_TAB,
        "sheet_url": f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit#gid=226918461",
        "generated_at": today,
        "source": "Flow Cockpit + Strategy Review col B",
        "project_count": len(projects),
        "with_growthpack_link": sum(1 for p in projects if p.get("growthpack_updated_link")),
        "projects": projects,
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / f"strategy_review_manifest_{today}.json"
    out.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"manifest: {out}")
    print(f"projetos: {len(projects)} | com GrowthPack Atualizado: {manifest['with_growthpack_link']}")


if __name__ == "__main__":
    main()
