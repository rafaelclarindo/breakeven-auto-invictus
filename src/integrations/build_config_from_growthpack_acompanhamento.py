#!/usr/bin/env python3
"""Monta config e-commerce da skill a partir do GP 3.0 (aba 6.0 Acompanhamento Mensal)."""
from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from statistics import median
from typing import Any

import openpyxl

PROJECT = Path(__file__).resolve().parents[2]
SHEET = "6.0 Acompanhamento Mensal"
ROWS = {
    "fee": 6,
    "media": 9,
    "sessions": 17,
    "orders": 27,
    "sales": 29,
    "revenue": 30,
}
MONTH_PT = {
    1: "Jan",
    2: "Fev",
    3: "Mar",
    4: "Abr",
    5: "Mai",
    6: "Jun",
    7: "Jul",
    8: "Ago",
    9: "Set",
    10: "Out",
    11: "Nov",
    12: "Dez",
}
# Referência 4.2 Funil — Mai/2026 a 15/Jun/2026 (Growth Pack O3NT)
FUNNEL_RATES = {
    "session_view": 2896 / 5131,
    "view_add": 311 / 2896,
    "add_view_cart": 213 / 311,
    "viewcart_checkout": 192 / 213,
    "checkout_shipping": 95 / 192,
    "shipping_payment": min(1.0, 96 / 95),
    "payment_order": 80 / 96,
    "order_sale": 65 / 80,
}


def num(value: Any, default: float = 0.0) -> float:
    try:
        return float(value) if value is not None else default
    except (TypeError, ValueError):
        return default


def safe_div(a: float, b: float, default: float = 0.0) -> float:
    return a / b if b else default


def month_label(dt: datetime) -> str:
    return f"{MONTH_PT[dt.month]}/{dt.year % 100:02d}"


def gradual(start: float, end: float, months: int = 7) -> list[float]:
    if months == 1:
        return [end]
    return [start + (end - start) * idx / (months - 1) for idx in range(months)]


def cap_rate(value: float, maximum: float = 0.95) -> float:
    return min(maximum, value)


def build_funnel_volumes(sessions: float, purchase: float) -> dict[str, float]:
    view_item = sessions * FUNNEL_RATES["session_view"]
    add_cart = view_item * FUNNEL_RATES["view_add"]
    view_cart = add_cart * FUNNEL_RATES["add_view_cart"]
    checkout = view_cart * FUNNEL_RATES["viewcart_checkout"]
    shipping = checkout * FUNNEL_RATES["checkout_shipping"]
    payment = shipping * FUNNEL_RATES["shipping_payment"]
    orders = payment * FUNNEL_RATES["payment_order"]
    modeled_purchase = orders * FUNNEL_RATES["order_sale"]
    if modeled_purchase > 0 and purchase > 0:
        scale = purchase / modeled_purchase
        add_cart *= scale
        view_cart *= scale
        checkout *= scale
        shipping *= scale
        payment *= scale
        orders *= scale
    return {
        "sessions": sessions,
        "view_item": view_item,
        "add_to_cart": add_cart,
        "view_cart": view_cart,
        "begin_checkout": checkout,
        "add_shipping_info": shipping,
        "add_payment_info": payment,
        "orders": orders,
        "purchase": purchase,
    }


def load_months(growthpack: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(growthpack, data_only=True)
    ws = wb[SHEET]
    months: list[dict[str, Any]] = []
    for col in range(2, ws.max_column + 1):
        dt = ws.cell(1, col).value
        if not isinstance(dt, datetime):
            continue
        item = {
            "label": month_label(dt),
            "fee": num(ws.cell(ROWS["fee"], col).value),
            "media": num(ws.cell(ROWS["media"], col).value),
            "sessions": num(ws.cell(ROWS["sessions"], col).value),
            "orders": num(ws.cell(ROWS["orders"], col).value),
            "sales": num(ws.cell(ROWS["sales"], col).value),
            "revenue": num(ws.cell(ROWS["revenue"], col).value),
        }
        if (
            item["fee"] > 0
            and item["media"] > 0
            and item["sessions"] > 0
            and item["sales"] > 0
            and item["revenue"] > 0
        ):
            months.append(item)
    if not months:
        raise ValueError(f"Nenhum mês válido em {SHEET}.")
    return months[-7:]


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera config e-commerce a partir do GP 3.0.")
    parser.add_argument("--project-folder", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--seasonal-context", type=str, default="")
    parser.add_argument("--seasonal-context-file", type=Path, default=None)
    args = parser.parse_args()

    project_folder = args.project_folder
    manifest = json.loads((project_folder / "source" / "manifest-entry.json").read_text(encoding="utf-8"))
    growthpack = project_folder / "source" / "growthpack.xlsx"
    months = load_months(growthpack)

    margin = num(manifest["margin_pct"]) / 100
    monthly_fee = num(manifest["fee"])
    monthly_media = num(manifest["media_planned"])
    current = months[-1]
    ticket = safe_div(sum(m["revenue"] for m in months), sum(m["sales"] for m in months), 357.65)

    session_view_med = median(
        cap_rate(safe_div(build_funnel_volumes(m["sessions"], m["sales"])["view_item"], m["sessions"]))
        for m in months
    )
    view_add_med = median(
        cap_rate(
            safe_div(
                build_funnel_volumes(m["sessions"], m["sales"])["add_to_cart"],
                build_funnel_volumes(m["sessions"], m["sales"])["view_item"],
            )
        )
        for m in months
    )
    view_cart_share_med = median(safe_div(m["orders"], m["sessions"]) for m in months)
    add_view_cart_med = median(
        cap_rate(
            safe_div(
                build_funnel_volumes(m["sessions"], m["sales"])["view_cart"],
                build_funnel_volumes(m["sessions"], m["sales"])["add_to_cart"],
            )
        )
        for m in months
    )
    order_sale_med = cap_rate(median(safe_div(m["sales"], m["orders"]) for m in months))

    source_months = [
        [m["label"], m["fee"], m["media"], m["sessions"], m["orders"], m["sales"], m["revenue"]]
        for m in months
    ]
    benchmark_months = []
    for m in months:
        funnel = build_funnel_volumes(m["sessions"], m["sales"])
        benchmark_months.append(
            [
                m["label"],
                funnel["sessions"],
                funnel["view_item"],
                funnel["add_to_cart"],
                funnel["view_cart"],
                funnel["begin_checkout"],
                funnel["add_shipping_info"],
                funnel["add_payment_info"],
                funnel["purchase"],
            ]
        )

    current_funnel_volumes = build_funnel_volumes(current["sessions"], current["sales"])
    min_cps = max(0.01, safe_div(current["media"], current["sessions"]) * 0.85)
    if args.seasonal_context_file:
        seasonal_context = args.seasonal_context_file.read_text(encoding="utf-8").strip()
    else:
        seasonal_context = (args.seasonal_context or "").strip()
    current_funnel = {
        **current_funnel_volumes,
        "page_view": current["sessions"] * 2.5,
        "orders": current["orders"],
        "sales": current["sales"],
        "revenue": current["revenue"],
        "media": current["media"],
    }

    breakeven_competence = (monthly_fee + monthly_media) / margin
    current_revenue = current["revenue"]
    realistic = [
        max(current_revenue * factor, breakeven_competence * 1.05)
        for factor in (1.08, 1.18, 1.32, 1.48, 1.62, 1.75, 1.88)
    ]

    scenarios = {
        "Pessimista": {
            "factor": (1.02, 1.06, 1.10, 1.14, 1.18, 1.22, 1.26),
            "tab_color": "#C55A11",
            "session_view_end": session_view_med * 0.95,
            "view_cart_end": view_cart_share_med * 0.92,
        },
        "Realista": {
            "revenue": realistic,
            "tab_color": "#5B9BD5",
            "session_view_end": session_view_med * 1.08,
            "view_cart_end": view_cart_share_med * 1.15,
        },
        "Otimista": {
            "factor": (1.12, 1.28, 1.45, 1.62, 1.78, 1.92, 2.05),
            "tab_color": "#70AD47",
            "session_view_end": session_view_med * 1.12,
            "view_cart_end": view_cart_share_med * 1.25,
        },
    }

    rendered_scenarios: dict[str, Any] = {}
    for name, scenario in scenarios.items():
        revenue = scenario.get("revenue") or [
            current_revenue * factor for factor in scenario["factor"]
        ]
        session_view = gradual(session_view_med * 0.98, scenario["session_view_end"])
        view_cart_share = gradual(view_cart_share_med * 0.95, scenario["view_cart_end"])
        rendered_scenarios[name] = {
            "media": [monthly_media] * 7,
            "revenue": [round(value, 2) for value in revenue],
            "ticket": [round(ticket * (1 + 0.008 * idx), 2) for idx in range(7)],
            "session_view": [cap_rate(value) for value in session_view],
            "view_cart_share": [max(0.0001, value) for value in view_cart_share],
            "add_view_cart": gradual(add_view_cart_med * 0.95, add_view_cart_med * 1.08),
            "viewcart_checkout": gradual(FUNNEL_RATES["viewcart_checkout"] * 0.98, cap_rate(FUNNEL_RATES["viewcart_checkout"] * 1.05)),
            "checkout_shipping": gradual(FUNNEL_RATES["checkout_shipping"] * 0.95, cap_rate(FUNNEL_RATES["checkout_shipping"] * 1.08)),
            "shipping_payment": gradual(FUNNEL_RATES["shipping_payment"] * 0.98, min(1.0, FUNNEL_RATES["shipping_payment"] * 1.02)),
            "payment_order": gradual(FUNNEL_RATES["payment_order"] * 0.98, cap_rate(FUNNEL_RATES["payment_order"] * 1.05)),
            "order_sale": gradual(order_sale_med * 0.98, cap_rate(order_sale_med * 1.05)),
            "tab_color": scenario["tab_color"],
        }

    config = {
        "client": manifest["name"],
        "project_model": "E-commerce D2C",
        "projection_rules": {
            "max_conversion_rate": 0.95,
            "min_cost_per_impression": min_cps,
            "media_lever_after_monthly_breakeven": True,
        },
        "current_period": f"{current['label']} fechado",
        "lt_period": f"{months[0]['label']} a {months[-1]['label']}",
        "margin": margin,
        "monthly_fee": monthly_fee,
        "monthly_media": monthly_media,
        "source_mapping": {
            "fee": f"Growth Pack > {SHEET} > linha {ROWS['fee']} (histórico); Flow/Strategy Review para competência",
            "media": f"Growth Pack > {SHEET} > linha {ROWS['media']} (realizado); Flow/Strategy Review para competência",
            "revenue": f"Growth Pack > {SHEET} > linha {ROWS['revenue']} (Faturamento)",
            "funnel_rates": "4.2 Funil — referência Mai/2026 (5131 sessões)",
            "sessions_orders_sales": f"Linhas {ROWS['sessions']}, {ROWS['orders']}, {ROWS['sales']}",
        },
        "source_months": source_months,
        "benchmark_months": benchmark_months,
        "current_funnel": current_funnel,
        "minimum_scenario": {
            "revenue": [round(value, 2) for value in realistic[:6]],
            "session_view": gradual(session_view_med, session_view_med * 1.05),
            "view_add": gradual(view_add_med * 0.98, cap_rate(view_add_med * 1.12)),
            "add_cart_purchase": cap_rate(
                safe_div(current["sales"], build_funnel_volumes(current["sessions"], current["sales"])["add_to_cart"])
            ),
            "approval_target": cap_rate(order_sale_med * 1.02),
            "order_sale": gradual(order_sale_med * 0.98, cap_rate(order_sale_med * 1.04)),
        },
        "scenarios": rendered_scenarios,
        "context": {
            "product": "Beleza e Cosméticos — e-commerce D2C",
            "phase": "ex_og_replan_180d — Strategy Review linha 70",
            "main_risk": "Faturamento abaixo do breakeven da competência; conversão e checkout",
            "seasonal": seasonal_context,
            "diagnosis": [
                f"LT de {len(months)} meses com fee e mídia preenchidos no Growth Pack ({months[0]['label']}–{months[-1]['label']}).",
                f"Breakeven da competência: R$ {breakeven_competence:,.2f} (fee R$ {monthly_fee:,.0f} + mídia R$ {monthly_media:,.0f} / margem {margin:.0%}).",
                f"Último mês fechado ({current['label']}): faturamento R$ {current_revenue:,.2f}, {current['sales']:.0f} vendas, {current['sessions']:.0f} sessões.",
                "Fee histórico no GP variou (R$ 8.000 → R$ 5.100); projeção usa fee/mídia/margem oficiais do Flow/Strategy Review.",
                "Funil intermediário estimado a partir das taxas da aba 4.2 Funil (Mai–Jun/2026).",
            ],
            "actions": [
                "Validar tracking GA4/WooCommerce nas etapas view_item → purchase",
                "Revisar frete, checkout e aprovação de pedidos com o time de e-commerce",
                "Confirmar se faturamento oficial usa linha 30 (Faturamento) vs WooCommerce puro",
            ],
        },
    }
    if seasonal_context:
        config["strategy_review_context"] = seasonal_context
        config["context"]["actions"].insert(
            0,
            "Alinhar campanhas e mix de produtos ao calendário sazonal descrito na Strategy Review (col. O).",
        )

    output = args.output or project_folder / "config.json"
    output.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

    gate = project_folder / "gate.md"
    gate.write_text(
        "\n".join(
            [
                "# Gate — O3NT",
                "",
                f"- Projeto: {manifest['name']}",
                "- Escopo: **E-commerce D2C**",
                f"- Growth Pack: `source/growthpack.xlsx` — aba `{SHEET}`",
                f"- LT: {config['lt_period']} ({len(months)} meses)",
                f"- Último mês: {config['current_period']}",
                f"- Fee competência: R$ {monthly_fee:,.2f}",
                f"- Mídia competência: R$ {monthly_media:,.2f}",
                f"- Margem: {margin:.0%}",
                f"- Breakeven competência: R$ {breakeven_competence:,.2f}",
                f"- Funil: taxas 4.2 Funil Mai/2026 aplicadas sobre sessões reais",
                "",
            ]
        ),
        encoding="utf-8",
    )
    print(output)


if __name__ == "__main__":
    main()
