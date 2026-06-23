#!/usr/bin/env python3
"""Resume abas e células-chave de um XLSX."""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl

KEYWORDS = [
    "fee",
    "mídia",
    "midia",
    "investimento",
    "faturamento",
    "receita",
    "sess",
    "session",
    "pedido",
    "venda",
    "purchase",
    "breakeven",
    "break-even",
    "margem",
]


def display(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def row_values(ws, row: int, start_col: int, end_col: int) -> list[Any]:
    return [ws.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]


def main() -> None:
    parser = argparse.ArgumentParser(description="Resume XLSX para mapeamento de breakeven.")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--max-matches", type=int, default=250)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=False, read_only=False)
    summary = {"file": str(args.xlsx), "sheets": [], "matches": []}

    pattern = re.compile("|".join(re.escape(k) for k in KEYWORDS), re.IGNORECASE)
    for ws in wb.worksheets:
        summary["sheets"].append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            }
        )
        for row in ws.iter_rows():
            for cell in row:
                value = display(cell.value)
                if not value or not pattern.search(value):
                    continue
                left = max(1, cell.column - 2)
                right = min(ws.max_column, cell.column + 12)
                summary["matches"].append(
                    {
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "value": value,
                        "row_values": row_values(ws, cell.row, left, right),
                    }
                )
                if len(summary["matches"]) >= args.max_matches:
                    break
            if len(summary["matches"]) >= args.max_matches:
                break
        if len(summary["matches"]) >= args.max_matches:
            break

    payload = json.dumps(summary, ensure_ascii=False, indent=2, default=str)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload + "\n", encoding="utf-8")
        print(args.output)
    else:
        print(payload)


if __name__ == "__main__":
    main()
