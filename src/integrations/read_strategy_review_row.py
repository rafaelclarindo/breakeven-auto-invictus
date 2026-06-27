#!/usr/bin/env python3
"""Lê uma linha da Strategy Review pelo nome do projeto."""
from __future__ import annotations

import argparse
import json
import unicodedata
from io import BytesIO
from pathlib import Path

import openpyxl
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

PROJECT = Path(__file__).resolve().parents[2]
SHEET_ID = "1MrUklD9tulNHsxWmAh3fcUUBlBdY-IHzSUuEPAz32YQ"
SHEET_TAB = "Start Strategy Review"
CRED_PATH = PROJECT.parent / "assessor-pessoal" / "mcp" / "credentials" / "google_sheets_token.json"


def norm(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(c for c in text if not unicodedata.combining(c))
    return " ".join(text.strip().lower().split())


def load_credentials() -> Credentials:
    info = json.loads(CRED_PATH.read_text(encoding="utf-8"))
    creds = Credentials(
        token=info.get("token"),
        refresh_token=info.get("refresh_token"),
        token_uri=info.get("token_uri", "https://oauth2.googleapis.com/token"),
        client_id=info.get("client_id"),
        client_secret=info.get("client_secret"),
        scopes=list(info.get("scopes") or []),
    )
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return creds


def cell_payload(cell) -> dict:
    return {
        "coordinate": cell.coordinate,
        "value": cell.value,
        "hyperlink": cell.hyperlink.target if cell.hyperlink else None,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Lê colunas da Strategy Review para um projeto.")
    parser.add_argument("project")
    args = parser.parse_args()

    drive = build("drive", "v3", credentials=load_credentials())
    xlsx = drive.files().export(
        fileId=SHEET_ID,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ).execute()
    wb = openpyxl.load_workbook(BytesIO(xlsx), read_only=False, data_only=False)
    ws = wb[SHEET_TAB] if SHEET_TAB in wb.sheetnames else wb.active

    needle = norm(args.project)
    for row in range(2, ws.max_row + 1):
        project_cell = ws[f"B{row}"]
        if needle in norm(project_cell.value) or norm(project_cell.value) in needle:
            # Layout Strategy Review (conferido 2026-06-24):
            # B=Projeto · H=LT · I=Fee · J=Mídia · K=Margem · L=MRR ·
            # M=Break-even antigo · N=GrowthPack · O=Retrospectiva · P=Sazonal.
            payload = {
                "row": row,
                "project": cell_payload(project_cell),
                "lt_h": cell_payload(ws[f"H{row}"]),
                "fee_i": cell_payload(ws[f"I{row}"]),
                "media_j": cell_payload(ws[f"J{row}"]),
                "margin_k": cell_payload(ws[f"K{row}"]),
                "mrr_l": cell_payload(ws[f"L{row}"]),
                "tm_recurrence_l": cell_payload(ws[f"L{row}"]),
                "old_breakeven_m": cell_payload(ws[f"M{row}"]),
                "growthpack_n": cell_payload(ws[f"N{row}"]),
                "retrospectiva_o": cell_payload(ws[f"O{row}"]),
                "seasonal_context_p": cell_payload(ws[f"P{row}"]),
            }
            print(json.dumps(payload, ensure_ascii=False, indent=2))
            return

    raise SystemExit(f"Projeto não encontrado: {args.project}")


if __name__ == "__main__":
    main()
