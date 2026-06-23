#!/usr/bin/env python3
"""Monta config da skill a partir da aba Tabela Mensal do Growth Pack."""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from statistics import median
from typing import Any

import openpyxl

PROJECT = Path(__file__).resolve().parents[2]


def num(value: Any, default: float = 0.0) -> float:
    try:
        return float(value) if value is not None else default
    except (TypeError, ValueError):
        return default


def safe_div(a: float, b: float, default: float = 0.0) -> float:
    return a / b if b else default


def gradual(start: float, end: float, months: int = 7) -> list[float]:
    if months == 1:
        return [end]
    return [start + (end - start) * idx / (months - 1) for idx in range(months)]


def month_label(month: str, year: Any) -> str:
    return f"{str(month).strip().title()}/{int(num(year))}"


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera config inicial da skill por Tabela Mensal.")
    parser.add_argument("--project-folder", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()

    project_folder = args.project_folder
    manifest = json.loads((project_folder / "source" / "manifest-entry.json").read_text(encoding="utf-8"))
    growthpack = project_folder / "source" / "growthpack.xlsx"
    wb = openpyxl.load_workbook(growthpack, data_only=True, read_only=False)
    ws = wb["Tabela Mensal"]

    rows = {
        "year": 1,
        "month": 3,
        "fee": 4,
        "ticket": 7,
        "media": 8,
        "clicks": 11,
        "sessions": 12,
        "add_cart": 13,
        "orders": 14,
        "sales": 15,
        "revenue": 16,
        "margin": 18,
    }

    valid = []
    for col in range(2, 8):
        fee = num(ws.cell(rows["fee"], col).value)
        media = num(ws.cell(rows["media"], col).value)
        revenue = num(ws.cell(rows["revenue"], col).value)
        if fee <= 0 or media <= 0 or revenue <= 0:
            continue
        month = month_label(ws.cell(rows["month"], col).value, ws.cell(rows["year"], col).value)
        sessions = num(ws.cell(rows["sessions"], col).value)
        add_cart = num(ws.cell(rows["add_cart"], col).value)
        orders = num(ws.cell(rows["orders"], col).value)
        sales = num(ws.cell(rows["sales"], col).value)
        valid.append(
            {
                "month": month,
                "fee": fee,
                "media": media,
                "clicks": num(ws.cell(rows["clicks"], col).value),
                "sessions": sessions,
                "add_cart": add_cart,
                "orders": orders,
                "sales": sales,
                "revenue": revenue,
                "ticket": num(ws.cell(rows["ticket"], col).value, safe_div(revenue, sales)),
                "gp_margin": num(ws.cell(rows["margin"], col).value),
            }
        )

    if not valid:
        raise ValueError("Nenhum mês válido encontrado na Tabela Mensal.")

    margin = num(manifest.get("margin_pct")) / 100
    monthly_fee = num(manifest.get("fee"))
    monthly_media = num(manifest.get("media_planned"))
    current = valid[-1]
    total_revenue = sum(row["revenue"] for row in valid)
    total_sales = sum(row["sales"] for row in valid)
    ticket = safe_div(total_revenue, total_sales, current["ticket"])

    order_rate = median(safe_div(row["orders"], row["sessions"]) for row in valid)
    add_to_order_rate = median(safe_div(row["orders"], row["add_cart"]) for row in valid)
    order_sale_rate = min(1.0, median(safe_div(row["sales"], row["orders"]) for row in valid))
    add_to_sale_rate = median(safe_div(row["sales"], row["add_cart"]) for row in valid)

    source_months = [
        [
            row["month"],
            row["fee"],
            row["media"],
            row["sessions"],
            row["orders"],
            row["sales"],
            row["revenue"],
        ]
        for row in valid
    ]

    # Mapeamento reduzido para inside sales dentro do contrato atual do gerador:
    # Sessões -> Add cart/leads -> Pedidos -> Vendas. Etapas intermediárias sem
    # fonte explícita no GP são mantidas iguais à etapa comprovada anterior.
    benchmark_months = [
        [
            row["month"],
            row["sessions"],
            row["sessions"],
            row["add_cart"],
            row["orders"],
            row["orders"],
            row["orders"],
            row["sales"],
            row["sales"],
        ]
        for row in valid
    ]

    current_funnel = {
        "sessions": current["sessions"],
        "page_view": current["clicks"],
        "view_item": current["sessions"],
        "add_to_cart": current["add_cart"],
        "view_cart": current["orders"],
        "begin_checkout": current["orders"],
        "add_shipping_info": current["orders"],
        "add_payment_info": current["sales"],
        "orders": current["orders"],
        "sales": current["sales"],
        "purchase": current["sales"],
        "revenue": current["revenue"],
        "media": current["media"],
    }

    current_revenue = current["revenue"]
    breakeven_competence = (monthly_fee + monthly_media) / margin
    realistic = [
        max(current_revenue * factor, breakeven_competence * 1.05)
        for factor in (1.00, 1.15, 1.30, 1.43, 1.54, 1.61, 1.68)
    ]
    scenarios = {
        "Pessimista": {
            "factor": (0.85, 0.90, 0.95, 1.00, 1.05, 1.10, 1.15),
            "tab_color": "#C55A11",
            "order_start": order_rate * 0.90,
            "order_end": order_rate * 1.05,
        },
        "Realista": {
            "revenue": realistic,
            "tab_color": "#5B9BD5",
            "order_start": order_rate,
            "order_end": order_rate * 1.25,
        },
        "Otimista": {
            "factor": (1.10, 1.25, 1.45, 1.60, 1.75, 1.90, 2.05),
            "tab_color": "#70AD47",
            "order_start": order_rate * 1.05,
            "order_end": order_rate * 1.45,
        },
    }

    rendered_scenarios = {}
    for name, scenario in scenarios.items():
        revenue = scenario.get("revenue") or [current_revenue * factor for factor in scenario["factor"]]
        view_cart_share = gradual(scenario["order_start"], scenario["order_end"])
        rendered_scenarios[name] = {
            "media": [monthly_media] * 7,
            "revenue": [round(value, 2) for value in revenue],
            "ticket": [round(ticket * (1 + 0.01 * idx), 2) for idx in range(7)],
            "session_view": [1.0] * 7,
            "view_cart_share": [max(0.0001, value) for value in view_cart_share],
            "add_view_cart": gradual(add_to_order_rate * 0.95, add_to_order_rate * 1.05),
            "viewcart_checkout": [1.0] * 7,
            "checkout_shipping": [1.0] * 7,
            "shipping_payment": [1.0] * 7,
            "payment_order": gradual(order_sale_rate * 0.98, min(1.0, order_sale_rate * 1.02)),
            "order_sale": gradual(order_sale_rate * 0.98, min(1.0, order_sale_rate * 1.02)),
            "tab_color": scenario["tab_color"],
        }

    config = {
        "client": manifest["name"],
        "project_model": "Inside Sales / geração de demanda",
        "current_period": f"{current['month']} fechado",
        "lt_period": f"{valid[0]['month']} a {valid[-1]['month']}",
        "margin": margin,
        "monthly_fee": monthly_fee,
        "monthly_media": monthly_media,
        "source_mapping": {
            "fee": "Growth Pack > Tabela Mensal > linha 4 para acumulado; Flow/Strategy Review para competência",
            "media": "Growth Pack > Tabela Mensal > linha 8 para acumulado; Flow/Strategy Review para competência",
            "revenue": "Growth Pack > Tabela Mensal > linha 16",
            "inside_sales_funnel": "Sessões linha 12, Add cart/leads linha 13, Pedidos linha 14, Vendas linha 15",
            "old_breakeven": "Strategy Review coluna M analisada em source/old-breakeven.xlsx",
        },
        "source_months": source_months,
        "benchmark_months": benchmark_months,
        "current_funnel": current_funnel,
        "minimum_scenario": {
            "revenue": [round(value, 2) for value in realistic[:6]],
            "session_view": [1.0] * 7,
            "view_add": gradual(median(safe_div(row["add_cart"], row["sessions"]) for row in valid), median(safe_div(row["add_cart"], row["sessions"]) for row in valid) * 1.15),
            "add_cart_purchase": add_to_sale_rate,
            "approval_target": min(1.0, order_sale_rate * 1.02),
            "order_sale": gradual(order_sale_rate * 0.98, min(1.0, order_sale_rate * 1.02)),
        },
        "scenarios": rendered_scenarios,
        "context": {
            "product": "Executar",
            "phase": "Teste Strategy Review / skill Jefferson",
            "main_risk": "Divergência entre Growth Pack, Flow e breakeven antigo; funil inside sales adaptado ao contrato e-commerce do gerador.",
            "diagnosis": [
                "O Growth Pack traz apenas Janeiro a Março/2024 com Fee V4 preenchido, então o LT usado no acumulado é de 3 meses.",
                "A competência usa Fee R$ 7.200, mídia R$ 30.000 e margem 15% da Strategy Review/Flow.",
                "O breakeven antigo da coluna M usa outro contrato financeiro na aba Projeção Break Even: Fee R$ 5.500 e mídia R$ 35.000.",
                "A aba Tabela Mensal do Growth Pack mostra margem 30%, mas a Strategy Review/Flow e o breakeven antigo usam 15% como margem de projeção.",
                "Como o projeto é inside sales, as etapas não disponíveis no contrato e-commerce foram mantidas como passagem neutra e documentadas no gate.",
            ],
            "actions": [
                "Confirmar se a margem oficial para Soma deve ser 15% ou 30%",
                "Confirmar se o Fee vigente para breakeven da competência é R$ 7.200 ou o Fee histórico de R$ 17.000 do Growth Pack",
                "Validar com o time se a etapa Add cart do Growth Pack representa leads/cadastros para Soma",
            ],
        },
    }

    output = args.output or project_folder / "config.json"
    output.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

    gate = project_folder / "gate.md"
    gate.write_text(
        "\n".join(
            [
                "# Gate — Soma Soluções",
                "",
                "- Projeto: SOMA SOLUCOES FINANCEIRAS LTDA",
                "- Escopo: inside sales / geração de demanda",
                "- Growth Pack: `source/growthpack.xlsx`",
                "- Breakeven antigo analisado: `source/old-breakeven.xlsx`",
                "- Linha Fee acumulado: `Tabela Mensal!A4:G4`",
                "- Linha mídia acumulada: `Tabela Mensal!A8:G8`",
                "- Linha faturamento: `Tabela Mensal!A16:G16`",
                "- Linhas funil: sessões `A12`, add cart/leads `A13`, pedidos `A14`, vendas `A15`",
                f"- Meses válidos/LT: {config['lt_period']}",
                f"- Último mês fechado: {config['current_period']}",
                f"- Fee competência: R$ {monthly_fee:,.2f} (Flow/Strategy Review)",
                f"- Mídia competência: R$ {monthly_media:,.2f} (Flow/Strategy Review)",
                f"- Margem usada: {margin:.2%} (Flow/Strategy Review)",
                "- Divergência: Growth Pack mostra margem 30% na `Tabela Mensal`, mas a Strategy Review e o breakeven antigo usam 15% para projeção.",
                "- Divergência: Growth Pack mostra Fee V4 de R$ 17.000 no acumulado; Strategy Review/Flow mostra Fee atual de R$ 7.200.",
                "- Adaptação: funil inside sales foi reduzido para caber no contrato atual do gerador sem inventar volumes intermediários.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    print(output)


if __name__ == "__main__":
    main()
