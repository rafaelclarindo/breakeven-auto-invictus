#!/usr/bin/env python3
"""Monta config inside sales Primeset (GP layout 6.0 Acompanhamento Mensal próprio)."""
from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from statistics import median
from typing import Any

import openpyxl

MAX_CONVERSION_RATE = 0.95
MIN_COST_PER_IMPRESSION = 0.01
MONTH_PT = {
    1: "Jan",
    2: "Fev",
    3: "Mar",
    4: "Abr",
    5: "Mai",
    6: "Jun",
    7: "Jul",
    8: "Ago",
    9: "Set",
    10: "Out",
    11: "Nov",
    12: "Dez",
}
SHEET = "6.0 Acompanhamento Mensal"
ROWS = {
    "date": 2,
    "media": 5,
    "impressions": 7,
    "clicks": 8,
    "leads": 9,
    "lead_quali": 10,
    "mqls": 13,
    "sqls": 15,
    "sales": 16,
    "revenue": 17,
}


def num(value: Any) -> float:
    try:
        return float(value) if value is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def div(a: float, b: float) -> float:
    return a / b if b else 0.0


def cap_rate(value: float) -> float:
    return min(MAX_CONVERSION_RATE, value)


def gradual(start: float, end: float, months: int = 7) -> list[float]:
    return [cap_rate(start + (end - start) * idx / (months - 1)) for idx in range(months)]


def month_label(dt: datetime) -> str:
    return f"{MONTH_PT[dt.month]}/{dt.year % 100:02d}"


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera config inside sales Primeset.")
    parser.add_argument("--project-folder", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument(
        "--lt-months",
        type=int,
        default=0,
        help="Limitar aos últimos N meses (0 = todos os meses válidos do GP).",
    )
    parser.add_argument("--seasonal-context", type=str, default="")
    args = parser.parse_args()

    project_folder = args.project_folder
    manifest = json.loads((project_folder / "source" / "manifest-entry.json").read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(project_folder / "source" / "growthpack.xlsx", data_only=True)
    ws = wb[SHEET]

    months: list[dict[str, Any]] = []
    monthly_fee = num(manifest["fee"])
    monthly_media = num(manifest["media_planned"])
    margin = num(manifest["margin_pct"]) / 100

    for col in range(2, ws.max_column + 1):
        dt = ws.cell(ROWS["date"], col).value
        if not isinstance(dt, datetime):
            continue
        item = {
            "label": month_label(dt),
            "fee": monthly_fee,
            "media": num(ws.cell(ROWS["media"], col).value),
            "impressions": num(ws.cell(ROWS["impressions"], col).value),
            "clicks": num(ws.cell(ROWS["clicks"], col).value),
            "leads": num(ws.cell(ROWS["leads"], col).value),
            "lead_quali": num(ws.cell(ROWS["lead_quali"], col).value),
            "mqls": num(ws.cell(ROWS["mqls"], col).value),
            "sqls": num(ws.cell(ROWS["sqls"], col).value),
            "sales": num(ws.cell(ROWS["sales"], col).value),
            "revenue": num(ws.cell(ROWS["revenue"], col).value),
        }
        if all(item[key] > 0 for key in ("media", "impressions", "clicks", "leads", "lead_quali", "mqls", "sqls", "sales", "revenue")):
            months.append(item)

    if not months:
        raise ValueError(f"Nenhum mês válido em {SHEET}.")

    if args.lt_months > 0:
        months = months[-args.lt_months :]
    ticket = div(sum(m["revenue"] for m in months), sum(m["sales"] for m in months))
    current = months[-1]
    current_cps = div(current["media"], current["impressions"])
    min_cps = max(MIN_COST_PER_IMPRESSION, current_cps * 0.85)

    rates = {
        "impression_click": median(cap_rate(div(m["clicks"], m["impressions"])) for m in months),
        "click_lead": median(cap_rate(div(m["leads"], m["clicks"])) for m in months),
        "lead_lead_quali": median(cap_rate(div(m["lead_quali"], m["leads"])) for m in months),
        "lead_quali_mql": median(cap_rate(div(m["mqls"], m["lead_quali"])) for m in months),
        "mql_sql": median(cap_rate(div(m["sqls"], m["mqls"])) for m in months),
        "sql_sale": median(cap_rate(div(m["sales"], m["sqls"])) for m in months),
    }
    current_rates = {
        "impression_click": cap_rate(div(current["clicks"], current["impressions"])),
        "click_lead": cap_rate(div(current["leads"], current["clicks"])),
        "lead_lead_quali": cap_rate(div(current["lead_quali"], current["leads"])),
        "lead_quali_mql": cap_rate(div(current["mqls"], current["lead_quali"])),
        "mql_sql": cap_rate(div(current["sqls"], current["mqls"])),
        "sql_sale": cap_rate(div(current["sales"], current["sqls"])),
    }

    def scenario_end(current_val: float, bench_val: float, multiplier: float) -> float:
        if multiplier < 1.0:
            return cap_rate(min(current_val * 0.98, bench_val * multiplier))
        return cap_rate(max(current_val * multiplier, bench_val * multiplier))

    def scenario_rate_series(rate_key: str, multiplier: float) -> list[float]:
        current_val = current_rates[rate_key]
        end_val = scenario_end(current_val, rates[rate_key], multiplier)
        return gradual(current_val, end_val)

    breakeven_competence = (monthly_fee + monthly_media) / margin
    current_revenue = current["revenue"]

    source_months = [
        [m["label"], m["fee"], m["media"], m["impressions"], m["sqls"], m["sales"], m["revenue"]]
        for m in months
    ]
    benchmark_months = [
        [
            m["label"],
            m["impressions"],
            m["clicks"],
            m["leads"],
            m["lead_quali"],
            m["mqls"],
            m["sqls"],
            m["sales"],
            m["sales"],
        ]
        for m in months
    ]
    current_funnel = {
        "sessions": current["impressions"],
        "page_view": current["impressions"],
        "view_item": current["clicks"],
        "add_to_cart": current["leads"],
        "view_cart": current["lead_quali"],
        "begin_checkout": current["mqls"],
        "add_shipping_info": current["sqls"],
        "add_payment_info": current["sales"],
        "orders": current["sqls"],
        "sales": current["sales"],
        "purchase": current["sales"],
        "revenue": current["revenue"],
        "media": current["media"],
    }

    def scenario(name: str, growth: float, multiplier: float, color: str) -> dict:
        revenue = []
        value = max(current_revenue, breakeven_competence * 1.02)
        for idx in range(7):
            if idx == 0:
                revenue.append(round(max(current_revenue * 1.05, breakeven_competence * 1.02), 2))
            else:
                value *= 1 + growth
                revenue.append(round(max(value, breakeven_competence * (1.02 + 0.03 * idx)), 2))
        imp_click = scenario_rate_series("impression_click", multiplier)
        click_lead = scenario_rate_series("click_lead", multiplier)
        lead_lead_quali = scenario_rate_series("lead_lead_quali", multiplier)
        lead_quali_mql = scenario_rate_series("lead_quali_mql", multiplier)
        mql_sql = scenario_rate_series("mql_sql", multiplier)
        sql_sale = scenario_rate_series("sql_sale", multiplier)
        return {
            "media": [monthly_media] * 7,
            "revenue": revenue,
            "ticket": [round(ticket * (1 + 0.005 * idx), 2) for idx in range(7)],
            "session_view": imp_click,
            "view_add": click_lead,
            "view_cart_share": [
                imp_click[i] * click_lead[i] * lead_lead_quali[i] * lead_quali_mql[i] for i in range(7)
            ],
            "add_view_cart": lead_lead_quali,
            "viewcart_checkout": lead_quali_mql,
            "checkout_shipping": mql_sql,
            "shipping_payment": sql_sale,
            "payment_order": [1.0] * 7,
            "order_sale": [1.0] * 7,
            "tab_color": color,
        }

    realista_scenario = scenario("Realista", 0.10, 1.05, "#5B9BD5")
    seasonal = (args.seasonal_context or "").strip()

    config = {
        "client": manifest["name"],
        "project_model": "Inside Sales",
        "funnel_has_lead_quali": True,
        "projection_rules": {
            "max_conversion_rate": MAX_CONVERSION_RATE,
            "min_cost_per_impression": min_cps,
            "media_lever_after_monthly_breakeven": True,
        },
        "current_period": f"{current['label']} fechado",
        "lt_period": f"{months[0]['label']} a {months[-1]['label']}",
        "margin": margin,
        "monthly_fee": monthly_fee,
        "monthly_media": monthly_media,
        "source_mapping": {
            "fee": "Strategy Review / Flow (fee R$ 15.836 — GP sem linha de fee mensal)",
            "media": f"Growth Pack > {SHEET} > linha 5 Investimento (histórico); Flow R$ 60.000 competência",
            "revenue": f"Growth Pack > {SHEET} > linha 17 Receita Faturada",
            "funnel": (
                "Impressões linha 7, Cliques linha 8, Leads linha 9, "
                "Leads no Ploomes linha 10, MQL linha 13, SQL linha 15, Vendas linha 16"
            ),
        },
        "source_months": source_months,
        "benchmark_months": benchmark_months,
        "current_funnel": current_funnel,
        "minimum_scenario": {
            "revenue": list(realista_scenario["revenue"]),
            "session_view": gradual(
                current_rates["impression_click"],
                cap_rate(max(current_rates["impression_click"] * 1.05, rates["impression_click"] * 1.05)),
            ),
            "view_add": gradual(
                current_rates["click_lead"],
                cap_rate(max(current_rates["click_lead"] * 1.05, rates["click_lead"] * 1.05)),
            ),
            "lead_lead_quali": gradual(
                current_rates["lead_lead_quali"],
                cap_rate(current_rates["lead_lead_quali"] * 1.03),
            ),
            "lead_quali_mql": gradual(
                current_rates["lead_quali_mql"],
                cap_rate(current_rates["lead_quali_mql"] * 1.08),
            ),
            "mql_sql": gradual(
                current_rates["mql_sql"],
                cap_rate(max(current_rates["mql_sql"] * 1.05, rates["mql_sql"] * 1.05)),
            ),
            "sql_sale": gradual(
                current_rates["sql_sale"],
                cap_rate(current_rates["sql_sale"] * 1.08),
            ),
            "add_cart_purchase": gradual(
                cap_rate(div(current["sales"], current["leads"])),
                cap_rate(div(current["sales"], current["leads"]) * 1.05),
            ),
            "approval_target": 1.0,
            "order_sale": [1.0] * 7,
        },
        "scenarios": {
            "Pessimista": scenario("Pessimista", 0.04, 0.96, "#C55A11"),
            "Realista": realista_scenario,
            "Otimista": scenario("Otimista", 0.16, 1.12, "#70AD47"),
        },
        "context": {
            "product": "Primeset — inside sales B2B (Ploomes)",
            "phase": "Strategy Review linha 45 — Growth Pack INSIDE SALES PRIMESET V1",
            "main_risk": "Receita faturada abaixo do breakeven da competência; queda de conversão SQL→venda em Jun/26.",
            "seasonal": seasonal or None,
            "diagnosis": [
                f"LT de {len(months)} meses no GP ({months[0]['label']}–{months[-1]['label']}) — todos os meses com funil completo.",
                f"Breakeven da competência: R$ {breakeven_competence:,.2f} (fee R$ {monthly_fee:,.0f} + mídia R$ {monthly_media:,.0f} / margem {margin:.0%}).",
                f"Último mês ({current['label']}): faturamento R$ {current_revenue:,.2f}, {current['sales']:.0f} vendas, {current['impressions']:.0f} impressões.",
                "Lead quali = Leads no Ploomes (linha 10); fee mensal V4 vem do Flow/SR (GP não registra fee).",
            ],
            "actions": [
                "Recuperar taxa SQL→venda e volume de MQLs com cadência comercial + qualificação no Ploomes",
                "Validar tracking Meta/Google vs Ploomes nas etapas lead quali → SQL",
            ],
        },
    }
    if seasonal:
        config["strategy_review_context"] = seasonal

    output = args.output or project_folder / "config.json"
    output.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

    gate = project_folder / "gate.md"
    gate.write_text(
        "\n".join(
            [
                "# Gate — Primeset (FABIO LUCHESI)",
                "",
                f"- Projeto: {manifest['name']}",
                "- Escopo: **Inside Sales** + Lead quali (Ploomes)",
                f"- Growth Pack: [INSIDE SALES PRIMESET V1]({manifest['growthpack_updated_link']})",
                f"- LT: {config['lt_period']} ({len(months)} meses)",
                f"- Fee competência: R$ {monthly_fee:,.2f}",
                f"- Mídia competência: R$ {monthly_media:,.2f}",
                f"- Margem: {margin:.0%}",
                f"- Breakeven competência: R$ {breakeven_competence:,.2f}",
                f"- Funil: impressões → cliques → leads → Ploomes → MQL → SQL → vendas",
                "",
            ]
        ),
        encoding="utf-8",
    )
    print(output)


if __name__ == "__main__":
    main()
