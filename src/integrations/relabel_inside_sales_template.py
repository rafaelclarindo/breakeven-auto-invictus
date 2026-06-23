#!/usr/bin/env python3
"""Renomeia labels do template da skill para funil inside sales mantendo a estrutura."""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

REPLACEMENTS = {
    "Breakeven E-commerce": "Breakeven Inside Sales",
    "Breakeven E-commerce —": "Breakeven Inside Sales —",
    "Projeção Breakeven E-commerce": "Projeção Breakeven Inside Sales",
    "Sessões": "Impressões",
    "Sessão → View item": "Impressões → Cliques",
    "View item": "Cliques",
    "View item → Add cart": "Cliques → Leads",
    "Taxa Sessão → View item": "Taxa Impressões → Cliques",
    "Taxa View item → Add to cart": "Taxa Cliques → Leads",
    "Add to cart": "Leads",
    "Add cart → View cart": "Leads → MQLs",
    "Taxa Add to cart → View cart": "Taxa Leads → MQLs",
    "View cart": "MQLs",
    "View cart → Checkout": "MQLs → SQLs",
    "Taxa View cart → Begin checkout": "Taxa MQLs → SQLs",
    "Begin checkout": "SQLs",
    "Shipping → Payment": "SQLs → Vendas",
    "Taxa Add shipping info → Add payment info": "Taxa SQLs → Vendas",
    "Sessão → Purchase": "Impressões → Vendas",
    "Taxa final Sessão → Venda": "Taxa final Leads → Vendas",
    "Taxa final Leads → Venda": "Taxa final Leads → Vendas",
    "Taxa de Conversão do Funil": "Taxa de Conversão Leads → Vendas",
    "Custo por sessão": "Custo por impressão",
    "Custo por pedido": "Custo por venda",
    "Taxa de Conversão Sessão → Pedidos": "Taxa Impressões → SQLs",
    "Taxa de Conversão Pedidos → Venda": "Taxa SQLs → Vendas",
    "Pedidos por mês": "SQLs por mês",
    "Sessões por mês": "Impressões por mês",
    "Custo por Sessão": "Custo por impressão",
    "Custo por Pedido": "Custo por SQL",
}

OLD_SHEET = "Breakeven 7M"
NEW_SHEET = "Breakeven"


def replace_text(value: str) -> str:
    output = value
    for old, new in REPLACEMENTS.items():
        output = output.replace(old, new)
    return output


def update_sheet_refs(workbook: openpyxl.Workbook) -> None:
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("=") and OLD_SHEET in value:
                    cell.value = value.replace(f"'{OLD_SHEET}'", f"'{NEW_SHEET}'")


def main() -> None:
    parser = argparse.ArgumentParser(description="Aplica labels inside sales no XLSX da skill.")
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input)
    for ws in wb.worksheets:
        if ws.title == OLD_SHEET:
            ws.title = NEW_SHEET
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and not cell.value.startswith("="):
                    cell.value = replace_text(cell.value)

    update_sheet_refs(wb)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(args.output)


if __name__ == "__main__":
    main()
