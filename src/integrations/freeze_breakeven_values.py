#!/usr/bin/env python3
"""Cria uma cópia estável do XLSX substituindo fórmulas por valores calculados."""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

ERRORS = {"#DIV/0!", "#REF!", "#VALUE!", "#NAME?", "#N/A"}


def is_error(value) -> bool:
    return isinstance(value, str) and any(error in value for error in ERRORS)


def main() -> None:
    parser = argparse.ArgumentParser(description="Congela fórmulas em valores para evitar erros no Google Sheets.")
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    formula_wb = openpyxl.load_workbook(args.input, data_only=False)
    value_wb = openpyxl.load_workbook(args.input, data_only=True)

    replaced = 0
    errors = []
    for ws in formula_wb.worksheets:
        value_ws = value_wb[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                cached = value_ws[cell.coordinate].value
                if is_error(cached):
                    errors.append((ws.title, cell.coordinate, cached))
                    cell.value = 0
                else:
                    cell.value = cached
                replaced += 1

    args.output.parent.mkdir(parents=True, exist_ok=True)
    formula_wb.save(args.output)
    print(f"replaced_formulas: {replaced}")
    print(f"cached_errors_replaced_with_zero: {len(errors)}")
    print(args.output)


if __name__ == "__main__":
    main()
